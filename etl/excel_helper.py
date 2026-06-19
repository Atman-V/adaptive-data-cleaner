"""
etl/excel_helper.py
Reads Excel files (.xlsx, .xlsm, .xls) into the same shape the rest of the
pipeline expects from a CSV: a list of headers + a list of dict rows
where every value is a plain string (mirroring csv.DictReader) — plus real
structural handling that plain cell-by-cell reading would miss:

  - Merged cells are unmerged by filling every cell in the merged range
    with the top-left value, so merged headers/data don't silently
    collapse into blanks.
  - A second header row (multi-level headers, e.g. "Q1" over "Revenue")
    is detected and folded into a single combined header.
  - Multiple stacked tables in one sheet (separated by fully-blank rows)
    are detected; the largest block is treated as the table to analyze.
"""

import os


def is_excel_file(filepath):
    return os.path.splitext(filepath)[1].lower() in (".xlsx", ".xlsm", ".xls")


def _cell_to_str(value):
    import datetime
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        if value.time() == datetime.time(0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _is_row_blank(row):
    return all(v is None or str(v).strip() == "" for v in row)


def read_excel_rows(filepath, sheet_name=None):
    """
    Reads the first sheet (or `sheet_name` if given) of an Excel workbook.

    Returns:
        (headers: list[str], rows: list[dict[str, str]], meta: dict) where
        meta has: merged_cells_filled, multi_level_header, tables_detected.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        # ── Unmerge: fill every cell in a merged range with its top-left value ──
        # openpyxl represents non-top-left cells in a merge as read-only
        # MergedCell objects, so the range must be unmerged first before any
        # cell in it can be assigned a value.
        merged_ranges = list(ws.merged_cells.ranges)
        merged_count = len(merged_ranges)
        for rng in merged_ranges:
            top_left_value = ws.cell(row=rng.min_row, column=rng.min_col).value
            ws.unmerge_cells(str(rng))
            for row in ws.iter_rows(min_row=rng.min_row, max_row=rng.max_row,
                                     min_col=rng.min_col, max_col=rng.max_col):
                for cell in row:
                    cell.value = top_left_value

        raw_rows = [[c.value for c in row] for row in ws.iter_rows()]
        meta = {"merged_cells_filled": merged_count, "multi_level_header": False, "tables_detected": 0}
        if not raw_rows:
            return [], [], meta

        # ── Header detection: fold a second header row into the first ──
        header_idx = 0
        header_row = raw_rows[0]
        multi_level = False
        if len(raw_rows) > 1:
            second = raw_rows[1]
            header_nonblank = sum(1 for v in header_row if v not in (None, ""))
            second_nonblank = sum(1 for v in second if v not in (None, ""))
            second_all_text = second_nonblank > 0 and all(
                v is None or isinstance(v, str) for v in second
            )
            # A second row that's mostly short text labels, immediately under
            # a header row that had merged cells, is a sub-header level.
            if second_all_text and merged_count > 0 and second_nonblank >= max(2, header_nonblank - 1):
                combined = []
                for top, sub in zip(header_row, second):
                    top_s = str(top).strip() if top not in (None, "") else ""
                    sub_s = str(sub).strip() if sub not in (None, "") else ""
                    if top_s and sub_s and top_s.lower() != sub_s.lower():
                        combined.append(f"{top_s} {sub_s}")
                    else:
                        combined.append(sub_s or top_s)
                header_row = combined
                header_idx = 1
                multi_level = True

        headers = [str(h).strip() if h not in (None, "") else f"col_{i+1}"
                   for i, h in enumerate(header_row)]
        body_rows = raw_rows[header_idx + 1:]

        # ── Multi-table detection: split on runs of fully-blank rows ──
        blocks, current = [], []
        for r in body_rows:
            if _is_row_blank(r):
                if current:
                    blocks.append(current)
                    current = []
            else:
                current.append(r)
        if current:
            blocks.append(current)

        meta["multi_level_header"] = multi_level
        meta["tables_detected"] = len(blocks)
        chosen = max(blocks, key=len) if blocks else []

        rows = []
        for raw_row in chosen:
            row = {}
            for i, h in enumerate(headers):
                v = raw_row[i] if i < len(raw_row) else None
                row[h] = _cell_to_str(v)
            rows.append(row)

        return headers, rows, meta
    finally:
        wb.close()
